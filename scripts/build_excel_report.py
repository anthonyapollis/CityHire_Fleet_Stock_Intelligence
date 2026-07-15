"""Build the CityHire Fleet & Stock Intelligence Excel workbook.

Raw source data is written as values (it is imported data, not a computed
result). Every KPI / summary cell is a live formula (SUMIFS/AVERAGEIFS/COUNTIFS/
IFERROR) referencing the raw sheets, so the workbook recalculates if the
underlying data changes.
"""
import csv
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.chart import LineChart, BarChart, Reference

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "data" / "raw"
PROC = ROOT / "data" / "processed"
OUT = ROOT / "excel" / "CityHire_Fleet_Stock_Intelligence.xlsx"

FONT_NAME = "Arial"
NAVY = "0C1114"
TEAL = "009B66"
AMBER = "D98E04"
RED = "C0392B"
GREEN = "009B66"
LIGHT_GREY = "F2F5F7"
WHITE = "FFFFFF"

header_font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
header_fill = PatternFill("solid", fgColor=NAVY)
title_font = Font(name=FONT_NAME, size=16, bold=True, color=NAVY)
subtitle_font = Font(name=FONT_NAME, size=10, italic=True, color="666666")
kpi_label_font = Font(name=FONT_NAME, size=10, bold=True, color="666666")
kpi_value_font = Font(name=FONT_NAME, size=20, bold=True, color=NAVY)
body_font = Font(name=FONT_NAME, size=10)
bold_font = Font(name=FONT_NAME, size=10, bold=True)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
wb.remove(wb.active)


def read_csv(name):
    path = RAW / f"{name}.csv" if (RAW / f"{name}.csv").exists() else PROC / f"{name}.csv"
    with open(path, encoding="utf-8") as f:
        r = csv.reader(f)
        rows = list(r)
    return rows[0], rows[1:]


def style_header(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def write_table(ws, headers, rows, start_row=1, start_col=1, number_cols=None, date_cols=None):
    number_cols = number_cols or set()
    date_cols = date_cols or set()
    for j, h in enumerate(headers):
        ws.cell(row=start_row, column=start_col + j, value=h)
    style_header(ws, start_row, len(headers), start_col)
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row):
            cell = ws.cell(row=start_row + i, column=start_col + j)
            if j in number_cols:
                try:
                    val = float(val)
                except (TypeError, ValueError):
                    pass
            cell.value = val
            cell.font = body_font
            cell.border = border
            if j in number_cols:
                cell.number_format = "#,##0.00"
            if (start_row + i) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
    return start_row + len(rows)


def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# =============================================================================
# SHEET: Read Me
# =============================================================================
ws = wb.create_sheet("Read Me")
ws["B2"] = "CityHire Fleet & Stock Intelligence"
ws["B2"].font = Font(name=FONT_NAME, size=20, bold=True, color=NAVY)
ws["B3"] = "A Data & Automation Analyst portfolio build — Anthony Apollis"
ws["B3"].font = subtitle_font

notes = [
    "",
    "PURPOSE",
    "Demonstration analytics build prepared for the 'Data & Automation Analyst' role (plant/tool "
    "hire sector), mapped directly against the job spec's 6-12 month success points and daily task list.",
    "",
    "DATA GROUNDING",
    "Category structure, subcategories and 220+ individual product/model names (e.g. 'Nifty Lift "
    "HR21 20.8m Hybrid Boom Lift', 'Hilti TE 3000-AVR Heavy Breaker') were extracted from the public "
    "product catalogue at cityhire.co.uk on 2026-07-15, across 7 categories scraped in depth (Plant, "
    "Powered Access, Breakers & Drills, Material Handling & Logistics, Lifting Equipment, Accommodation "
    "Storage & Welfare, Cutting & Grinding) and extrapolated in the same style for the remaining 17 "
    "listed categories to cover the full fleet range.",
    "",
    "IMPORTANT: all fleet quantities, serial/stock codes, purchase dates, costs, utilisation figures, "
    "workshop status, repair jobs, approvals and stock transfers below are SYNTHETIC data generated for "
    "this demo — they are not City Hire's real fleet, financial or operational records.",
    "",
    "HOW TO READ THIS WORKBOOK",
    "1. Dashboard — the 3 headline success metrics + monthly trend chart, computed live via formulas.",
    "2. Stock Master — 1,612-asset fleet register (the source table every other sheet references).",
    "3. Capex vs Budget — monthly spend vs budget by category, with an exception flag for overspend.",
    "4. ROI Report — estimated annual revenue vs capex by category, and the lowest-ROI assets to review.",
    "5. Approvals — auction/parts/capital requests >£500 awaiting sign-off.",
    "6. External Repairs — supplier repair jobs and which need approval before proceeding.",
    "7. Stock Shortage & Transfers — fleeting activity between depots, and unfulfilled shortage requests.",
    "8. Min Stock Alerts — asset categories currently trading below their minimum stock level at a depot.",
    "",
    "All KPI and summary cells use live formulas (SUMIFS / AVERAGEIFS / COUNTIFS / IFERROR) against the "
    "raw data sheets — change the underlying data and the workbook recalculates.",
]
r = 5
for line in notes:
    cell = ws.cell(row=r, column=2, value=line)
    if line.isupper() and line:
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color=TEAL)
    else:
        cell.font = body_font
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 28 if len(line) > 90 else 15
    r += 1
autosize(ws, {"A": 3, "B": 110})
ws.sheet_view.showGridLines = False

# =============================================================================
# SHEET: Stock Master (raw)
# =============================================================================
ws = wb.create_sheet("Stock Master")
headers, rows = read_csv("stock_master")
num_idx = {headers.index(c) for c in ["capex_gbp", "book_value_gbp", "age_years", "utilisation_pct_ytd",
                                       "day_rate_gbp", "min_stock_level", "current_units_at_depot"]}
last_row = write_table(ws, headers, rows, number_cols=num_idx)
ws.freeze_panes = "A2"
autosize(ws, {get_column_letter(i + 1): 16 for i in range(len(headers))})
STOCK_ROWS = len(rows)
STOCK_LAST_ROW = last_row

# conditional formatting: utilisation column red/amber/green
util_col = headers.index("utilisation_pct_ytd") + 1
util_letter = get_column_letter(util_col)
ws.conditional_formatting.add(
    f"{util_letter}2:{util_letter}{STOCK_LAST_ROW}",
    ColorScaleRule(start_type="min", start_color="C0392B", mid_type="num", mid_value=60,
                    mid_color="D98E04", end_type="max", end_color="1E824C"),
)
status_col = get_column_letter(headers.index("status") + 1)
ws.conditional_formatting.add(
    f"{status_col}2:{status_col}{STOCK_LAST_ROW}",
    CellIsRule(operator="equal", formula=['"Workshop"'], fill=PatternFill("solid", fgColor="FADBD8")),
)
ws.conditional_formatting.add(
    f"{status_col}2:{status_col}{STOCK_LAST_ROW}",
    CellIsRule(operator="equal", formula=['"Auction Pending"'], fill=PatternFill("solid", fgColor="FCF3CF")),
)

print(f"Stock Master: {STOCK_ROWS} rows, last_row={STOCK_LAST_ROW}")

# =============================================================================
# SHEET: Monthly Trend (raw, small — drives dashboard chart)
# =============================================================================
ws = wb.create_sheet("Monthly Trend")
headers, rows = read_csv("monthly_fleet_trend")
num_idx = {headers.index(c) for c in ["fleet_units", "units_on_hire", "units_in_workshop",
                                       "utilisation_pct", "workshop_pct"]}
last_row = write_table(ws, headers, rows, number_cols=num_idx)
autosize(ws, {get_column_letter(i + 1): 16 for i in range(len(headers))})
TREND_LAST_ROW = last_row
TREND_UTIL_COL = get_column_letter(headers.index("utilisation_pct") + 1)
TREND_WORKSHOP_COL = get_column_letter(headers.index("workshop_pct") + 1)
TREND_MONTH_COL = get_column_letter(headers.index("month") + 1)

# =============================================================================
# SHEET: Capex vs Budget (raw + category summary via formulas)
# =============================================================================
ws = wb.create_sheet("Capex vs Budget")
ws["A1"] = "Category Summary (YTD) — live SUMIFS against 'Capex Raw' rows below"
ws["A1"].font = bold_font
cat_headers = ["category", "ytd_budget_gbp", "ytd_actual_gbp", "ytd_variance_gbp", "pct_of_budget_used", "status"]
for j, h in enumerate(cat_headers):
    ws.cell(row=2, column=j + 1, value=h)
style_header(ws, 2, len(cat_headers))

raw_headers, raw_rows = read_csv("capex_monthly")
RAW_START = 2 + 1 + len(set(r[1] for r in raw_rows)) + 3  # placeholder, fixed below
categories_sorted = sorted(set(r[1] for r in raw_rows))
n_cats = len(categories_sorted)
summary_last_row = 2 + n_cats
for i, cat in enumerate(categories_sorted):
    row = 3 + i
    ws.cell(row=row, column=1, value=cat)

RAW_TABLE_START = summary_last_row + 3
ws.cell(row=RAW_TABLE_START - 1, column=1, value="Capex Raw (monthly by category)").font = bold_font
raw_num_idx = {raw_headers.index(c) for c in ["budget_gbp", "actual_spend_gbp", "variance_gbp"]}
raw_last_row = write_table(ws, raw_headers, raw_rows, start_row=RAW_TABLE_START, number_cols=raw_num_idx)
cat_col = get_column_letter(raw_headers.index("category") + 1)
budget_col = get_column_letter(raw_headers.index("budget_gbp") + 1)
actual_col = get_column_letter(raw_headers.index("actual_spend_gbp") + 1)
raw_data_first = RAW_TABLE_START + 1
raw_data_last = raw_last_row

for i, cat in enumerate(categories_sorted):
    row = 3 + i
    ws.cell(row=row, column=2, value=(
        f"=SUMIFS('Capex vs Budget'!{budget_col}{raw_data_first}:{budget_col}{raw_data_last},"
        f"'Capex vs Budget'!{cat_col}{raw_data_first}:{cat_col}{raw_data_last},A{row})"
    ))
    ws.cell(row=row, column=3, value=(
        f"=SUMIFS('Capex vs Budget'!{actual_col}{raw_data_first}:{actual_col}{raw_data_last},"
        f"'Capex vs Budget'!{cat_col}{raw_data_first}:{cat_col}{raw_data_last},A{row})"
    ))
    ws.cell(row=row, column=4, value=f"=C{row}-B{row}")
    ws.cell(row=row, column=5, value=f"=IFERROR(C{row}/B{row},0)")
    ws.cell(row=row, column=6, value=f'=IF(D{row}>0,"OVER BUDGET","Within budget")')
    for c in range(1, 7):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.border = border
    ws.cell(row=row, column=2).number_format = "#,##0"
    ws.cell(row=row, column=3).number_format = "#,##0"
    ws.cell(row=row, column=4).number_format = "#,##0;(#,##0)"
    ws.cell(row=row, column=5).number_format = "0.0%"

ws.conditional_formatting.add(
    f"F3:F{summary_last_row}",
    CellIsRule(operator="equal", formula=['"OVER BUDGET"'], fill=PatternFill("solid", fgColor="FADBD8"),
               font=Font(name=FONT_NAME, color=RED, bold=True)),
)
autosize(ws, {"A": 30, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 14, "H": 14, "I": 16, "J": 14})
CAPEX_SUMMARY_LAST = summary_last_row
CAPEX_CAT_RANGE = f"'Capex vs Budget'!A3:A{summary_last_row}"
CAPEX_VAR_RANGE = f"'Capex vs Budget'!D3:D{summary_last_row}"

# =============================================================================
# SHEET: ROI Report (category summary via formulas + lowest-ROI asset list)
# =============================================================================
ws = wb.create_sheet("ROI Report")
ws["A1"] = "Category ROI Summary — live formulas against 'Stock Master'"
ws["A1"].font = bold_font
roi_headers = ["category", "units", "total_capex_gbp", "avg_utilisation_pct",
               "est_annual_revenue_gbp", "est_roi_pct"]
for j, h in enumerate(roi_headers):
    ws.cell(row=2, column=j + 1, value=h)
style_header(ws, 2, len(roi_headers))

stock_headers, _ = read_csv("stock_master")
sm_cat_col = get_column_letter(stock_headers.index("category") + 1)
sm_capex_col = get_column_letter(stock_headers.index("capex_gbp") + 1)
sm_util_col = get_column_letter(stock_headers.index("utilisation_pct_ytd") + 1)
sm_rate_col = get_column_letter(stock_headers.index("day_rate_gbp") + 1)
sm_first, sm_last = 2, STOCK_LAST_ROW

roi_categories = sorted(set(r[stock_headers.index("category")] for r in rows if False)) or categories_sorted
for i, cat in enumerate(roi_categories):
    row = 3 + i
    ws.cell(row=row, column=1, value=cat)
    ws.cell(row=row, column=2, value=(
        f"=COUNTIF('Stock Master'!{sm_cat_col}{sm_first}:{sm_cat_col}{sm_last},A{row})"
    ))
    ws.cell(row=row, column=3, value=(
        f"=SUMIF('Stock Master'!{sm_cat_col}{sm_first}:{sm_cat_col}{sm_last},A{row},"
        f"'Stock Master'!{sm_capex_col}{sm_first}:{sm_capex_col}{sm_last})"
    ))
    ws.cell(row=row, column=4, value=(
        f"=AVERAGEIF('Stock Master'!{sm_cat_col}{sm_first}:{sm_cat_col}{sm_last},A{row},"
        f"'Stock Master'!{sm_util_col}{sm_first}:{sm_util_col}{sm_last})"
    ))
    ws.cell(row=row, column=5, value=(
        f"=SUMPRODUCT(('Stock Master'!{sm_cat_col}{sm_first}:{sm_cat_col}{sm_last}=A{row})*"
        f"'Stock Master'!{sm_rate_col}{sm_first}:{sm_rate_col}{sm_last}*"
        f"'Stock Master'!{sm_util_col}{sm_first}:{sm_util_col}{sm_last}/100*365)"
    ))
    ws.cell(row=row, column=6, value=f"=IFERROR(E{row}/C{row},0)")
    for c in range(1, 7):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.border = border
    ws.cell(row=row, column=3).number_format = "#,##0"
    ws.cell(row=row, column=4).number_format = "0.0"
    ws.cell(row=row, column=5).number_format = "#,##0"
    ws.cell(row=row, column=6).number_format = "0.0%"

roi_summary_last = 2 + len(roi_categories)
ws.conditional_formatting.add(
    f"F3:F{roi_summary_last}",
    ColorScaleRule(start_type="min", start_color="C0392B", mid_type="percentile", mid_value=50,
                    mid_color="D98E04", end_type="max", end_color="1E824C"),
)

low_roi_start = roi_summary_last + 3
ws.cell(row=low_roi_start - 1, column=1, value="Lowest-ROI Individual Assets (auction/disposal review list)").font = bold_font
low_headers, low_rows = read_csv("lowest_roi_assets")
low_num_idx = {low_headers.index(c) for c in ["capex_gbp", "book_value_gbp", "utilisation_pct_ytd",
                                               "est_annual_revenue_gbp", "est_roi_pct"]}
write_table(ws, low_headers, low_rows, start_row=low_start if (low_start := low_roi_start) else low_roi_start,
            number_cols=low_num_idx)
autosize(ws, {get_column_letter(i + 1): 18 for i in range(9)})
ROI_CAT_RANGE = f"'ROI Report'!A3:A{roi_summary_last}"
ROI_PCT_RANGE = f"'ROI Report'!F3:F{roi_summary_last}"

# =============================================================================
# SHEET: Approvals
# =============================================================================
ws = wb.create_sheet("Approvals")
ws["A1"] = "Approvals Summary by Request Type — live formulas"
ws["A1"].font = bold_font
sum_headers = ["request_type", "requests", "total_value_gbp", "avg_value_gbp", "approved", "pending", "rejected"]
for j, h in enumerate(sum_headers):
    ws.cell(row=2, column=j + 1, value=h)
style_header(ws, 2, len(sum_headers))

app_headers, app_rows = read_csv("approvals")
apr_type_col = get_column_letter(app_headers.index("request_type") + 1)
apr_val_col = get_column_letter(app_headers.index("value_gbp") + 1)
apr_status_col = get_column_letter(app_headers.index("status") + 1)
req_types = sorted(set(r[app_headers.index("request_type")] for r in app_rows))

pending_start = 3 + len(req_types) + 3
raw_first = pending_start + 2
raw_last = raw_first + len(app_rows) - 1

for i, rt in enumerate(req_types):
    row = 3 + i
    ws.cell(row=row, column=1, value=rt)
    ws.cell(row=row, column=2, value=f"=COUNTIF({apr_type_col}{raw_first}:{apr_type_col}{raw_last},A{row})")
    ws.cell(row=row, column=3, value=f"=SUMIF({apr_type_col}{raw_first}:{apr_type_col}{raw_last},A{row},{apr_val_col}{raw_first}:{apr_val_col}{raw_last})")
    ws.cell(row=row, column=4, value=f"=IFERROR(C{row}/B{row},0)")
    ws.cell(row=row, column=5, value=f'=COUNTIFS({apr_type_col}{raw_first}:{apr_type_col}{raw_last},A{row},{apr_status_col}{raw_first}:{apr_status_col}{raw_last},"Approved")')
    ws.cell(row=row, column=6, value=f'=COUNTIFS({apr_type_col}{raw_first}:{apr_type_col}{raw_last},A{row},{apr_status_col}{raw_first}:{apr_status_col}{raw_last},"Pending")')
    ws.cell(row=row, column=7, value=f'=COUNTIFS({apr_type_col}{raw_first}:{apr_type_col}{raw_last},A{row},{apr_status_col}{raw_first}:{apr_status_col}{raw_last},"Rejected")')
    for c in range(1, 8):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.border = border
    ws.cell(row=row, column=3).number_format = "#,##0"
    ws.cell(row=row, column=4).number_format = "#,##0"

ws.cell(row=pending_start - 1, column=1, value="Full Approvals Log (all requests >£500 route here)").font = bold_font
app_num_idx = {app_headers.index("value_gbp")}
write_table(ws, app_headers, app_rows, start_row=pending_start, number_cols=app_num_idx)
pending_col = get_column_letter(app_headers.index("status") + 1)
ws.conditional_formatting.add(
    f"{pending_col}{raw_first}:{pending_col}{raw_last}",
    CellIsRule(operator="equal", formula=['"Pending"'], fill=PatternFill("solid", fgColor="FCF3CF")),
)
autosize(ws, {get_column_letter(i + 1): 18 for i in range(8)})
APPROVALS_PENDING_TOTAL_ROW = 3 + len(req_types)

# =============================================================================
# SHEET: External Repairs
# =============================================================================
ws = wb.create_sheet("External Repairs")
ws["A1"] = "Repairs by Supplier — live formulas"
ws["A1"].font = bold_font
rep_sum_headers = ["supplier", "repair_jobs", "total_cost_gbp", "avg_days_out_of_service", "pending_approvals"]
for j, h in enumerate(rep_sum_headers):
    ws.cell(row=2, column=j + 1, value=h)
style_header(ws, 2, len(rep_sum_headers))

rep_headers, rep_rows = read_csv("external_repairs")
sup_col = get_column_letter(rep_headers.index("supplier") + 1)
cost_col = get_column_letter(rep_headers.index("cost_gbp") + 1)
days_col = get_column_letter(rep_headers.index("days_out_of_service") + 1)
req_appr_col = get_column_letter(rep_headers.index("requires_approval") + 1)
appr_status_col = get_column_letter(rep_headers.index("approval_status") + 1)
suppliers = sorted(set(r[rep_headers.index("supplier")] for r in rep_rows))

rep_raw_start = 3 + len(suppliers) + 3
rep_raw_first = rep_raw_start + 1
rep_raw_last = rep_raw_first + len(rep_rows) - 1

for i, sup in enumerate(suppliers):
    row = 3 + i
    ws.cell(row=row, column=1, value=sup)
    ws.cell(row=row, column=2, value=f"=COUNTIF({sup_col}{rep_raw_first}:{sup_col}{rep_raw_last},A{row})")
    ws.cell(row=row, column=3, value=f"=SUMIF({sup_col}{rep_raw_first}:{sup_col}{rep_raw_last},A{row},{cost_col}{rep_raw_first}:{cost_col}{rep_raw_last})")
    ws.cell(row=row, column=4, value=f"=AVERAGEIF({sup_col}{rep_raw_first}:{sup_col}{rep_raw_last},A{row},{days_col}{rep_raw_first}:{days_col}{rep_raw_last})")
    ws.cell(row=row, column=5, value=(
        f'=COUNTIFS({sup_col}{rep_raw_first}:{sup_col}{rep_raw_last},A{row},'
        f'{appr_status_col}{rep_raw_first}:{appr_status_col}{rep_raw_last},"Pending")'
    ))
    for c in range(1, 6):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.border = border
    ws.cell(row=row, column=3).number_format = "#,##0"
    ws.cell(row=row, column=4).number_format = "0.0"

ws.cell(row=rep_raw_start, column=1, value="Full External Repairs Log").font = bold_font
rep_num_idx = {rep_headers.index(c) for c in ["cost_gbp", "days_out_of_service"]}
write_table(ws, rep_headers, rep_rows, start_row=rep_raw_first - 1, number_cols=rep_num_idx)
autosize(ws, {get_column_letter(i + 1): 20 for i in range(9)})

# =============================================================================
# SHEET: Stock Shortage & Transfers
# =============================================================================
ws = wb.create_sheet("Stock Shortage")
xf_headers, xf_rows = read_csv("stock_transfers")
fulfilled_idx = xf_headers.index("fulfilled")
xf_rows = [[*(v if i != fulfilled_idx else (v == "True") for i, v in enumerate(row))] for row in xf_rows]
ws["A1"] = "Depot-to-Depot Fleeting Log (stock shortage balancing)"
ws["A1"].font = bold_font
xf_num_idx = {xf_headers.index("units_transferred")}
last_row = write_table(ws, xf_headers, xf_rows, start_row=2, number_cols=xf_num_idx)
fulfilled_col = get_column_letter(xf_headers.index("fulfilled") + 1)
ws.conditional_formatting.add(
    f"{fulfilled_col}3:{fulfilled_col}{last_row}",
    CellIsRule(operator="equal", formula=["FALSE"], fill=PatternFill("solid", fgColor="FADBD8")),
)
autosize(ws, {get_column_letter(i + 1): 20 for i in range(len(xf_headers))})
XFER_LAST_ROW = last_row
XFER_FULFILLED_COL = fulfilled_col

# =============================================================================
# SHEET: Min Stock Alerts
# =============================================================================
ws = wb.create_sheet("Min Stock Alerts")
ms_headers, ms_rows = read_csv("stock_below_minimum")
ws["A1"] = "Assets Currently Below Minimum Stock Level, by Depot"
ws["A1"].font = bold_font
ms_num_idx = {i for i, h in enumerate(ms_headers) if h in
              ("assets_of_type", "units_available", "min_stock_level")}
last_row = write_table(ws, ms_headers, ms_rows, start_row=2, number_cols=ms_num_idx)
autosize(ws, {get_column_letter(i + 1): 20 for i in range(len(ms_headers))})

# =============================================================================
# SHEET: Dashboard (build last so all reference sheets/rows exist)
# =============================================================================
dash = wb.create_sheet("Dashboard", 0)
dash.sheet_view.showGridLines = False
dash["B1"] = "CityHire Fleet & Stock Intelligence — Executive Dashboard"
dash["B1"].font = title_font
dash["B2"] = "Live workbook — all figures below are formulas referencing the data sheets"
dash["B2"].font = subtitle_font
dash.merge_cells("B1:K1")
dash.merge_cells("B2:K2")

kpi_defs = [
    ("Fleet In-Service % (latest month)", f"=1-'Monthly Trend'!{TREND_WORKSHOP_COL}{TREND_LAST_ROW}/100",
     "0.0%", "Target: >= 85% (workshop <= 15%)"),
    ("Fleet Utilisation % (latest month)", f"='Monthly Trend'!{TREND_UTIL_COL}{TREND_LAST_ROW}/100",
     "0.0%", "Target: >= 60%"),
    ("Categories Over Capex Budget (YTD)", f'=COUNTIF({CAPEX_CAT_RANGE.replace("A3", "F3").replace("A"+str(CAPEX_SUMMARY_LAST),"F"+str(CAPEX_SUMMARY_LAST))},"OVER BUDGET")',
     "0", "Target: 0"),
]
# fix the third formula reference cleanly
kpi_defs[2] = ("Categories Over Capex Budget (YTD)",
               f"=COUNTIF('Capex vs Budget'!F3:F{CAPEX_SUMMARY_LAST},\"OVER BUDGET\")",
               "0", "Target: 0")

col = 2
for label, formula, fmt, note in kpi_defs:
    c0 = get_column_letter(col)
    c1 = get_column_letter(col + 2)
    dash.merge_cells(f"{c0}4:{c1}4")
    dash.merge_cells(f"{c0}5:{c1}6")
    dash.merge_cells(f"{c0}7:{c1}7")
    dash[f"{c0}4"] = label
    dash[f"{c0}4"].font = kpi_label_font
    dash[f"{c0}5"] = formula
    dash[f"{c0}5"].font = kpi_value_font
    dash[f"{c0}5"].number_format = fmt
    dash[f"{c0}5"].alignment = Alignment(horizontal="left", vertical="center")
    dash[f"{c0}7"] = note
    dash[f"{c0}7"].font = Font(name=FONT_NAME, size=9, italic=True, color="888888")
    for rr in range(4, 8):
        for cc in range(col, col + 3):
            dash.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor=LIGHT_GREY)
    col += 4

# second-row KPIs: pending approvals value, pending repairs, unfulfilled transfers
kpi_defs2 = [
    ("Approvals Pending (>£500)", f"=COUNTIF('Approvals'!G{pending_start}:G{raw_last},\"Pending\")", "0",
     "Awaiting Ops & Systems Manager sign-off"),
    ("Assets Below Min Stock Level", f"=COUNTA('Min Stock Alerts'!A3:A{last_row if False else 0})", "0", ""),
    ("Unfulfilled Depot Transfers",
     f"=COUNTIF('Stock Shortage'!{XFER_FULFILLED_COL}3:{XFER_FULFILLED_COL}{XFER_LAST_ROW},FALSE)", "0",
     "Shortage requests still open"),
]
ms_last_row = 2 + len(ms_rows)
kpi_defs2[1] = ("Assets Below Min Stock Level", f"=COUNTA('Min Stock Alerts'!A3:A{ms_last_row})", "0",
                "Category/depot combinations short of minimum holding")

col = 2
for label, formula, fmt, note in kpi_defs2:
    c0 = get_column_letter(col)
    c1 = get_column_letter(col + 2)
    dash.merge_cells(f"{c0}9:{c1}9")
    dash.merge_cells(f"{c0}10:{c1}11")
    dash.merge_cells(f"{c0}12:{c1}12")
    dash[f"{c0}9"] = label
    dash[f"{c0}9"].font = kpi_label_font
    dash[f"{c0}10"] = formula
    dash[f"{c0}10"].font = kpi_value_font
    dash[f"{c0}10"].number_format = fmt
    dash[f"{c0}12"] = note
    dash[f"{c0}12"].font = Font(name=FONT_NAME, size=9, italic=True, color="888888")
    for rr in range(9, 13):
        for cc in range(col, col + 3):
            dash.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor=LIGHT_GREY)
    col += 4

# Monthly trend chart
chart = LineChart()
chart.title = "12-Month Trend: Utilisation % vs Workshop (In-Service Target)"
chart.style = 10
chart.y_axis.title = "%"
chart.x_axis.title = "Month"
util_ref = Reference(wb["Monthly Trend"], min_col=headers.index("utilisation_pct") + 1,
                      min_row=1, max_row=TREND_LAST_ROW)
ws_headers_trend, _ = read_csv("monthly_fleet_trend")
util_col_idx = ws_headers_trend.index("utilisation_pct") + 1
workshop_col_idx = ws_headers_trend.index("workshop_pct") + 1
month_col_idx = ws_headers_trend.index("month") + 1
util_ref = Reference(wb["Monthly Trend"], min_col=util_col_idx, min_row=1, max_row=TREND_LAST_ROW)
workshop_ref = Reference(wb["Monthly Trend"], min_col=workshop_col_idx, min_row=1, max_row=TREND_LAST_ROW)
cats_ref = Reference(wb["Monthly Trend"], min_col=month_col_idx, min_row=2, max_row=TREND_LAST_ROW)
chart.add_data(util_ref, titles_from_data=True)
chart.add_data(workshop_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.width = 24
chart.height = 11
dash.add_chart(chart, "B15")

# Capex over/under budget bar chart
bar = BarChart()
bar.title = "YTD Capex Variance by Category (£, actual - budget)"
bar.style = 12
bar.y_axis.title = "£ variance"
var_ref = Reference(wb["Capex vs Budget"], min_col=4, min_row=2, max_row=CAPEX_SUMMARY_LAST)
cat_ref = Reference(wb["Capex vs Budget"], min_col=1, min_row=3, max_row=CAPEX_SUMMARY_LAST)
bar.add_data(var_ref, titles_from_data=True)
bar.set_categories(cat_ref)
bar.width = 24
bar.height = 14
dash.add_chart(bar, "L15")

autosize(dash, {get_column_letter(i): 12 for i in range(1, 14)})
dash.column_dimensions["A"].width = 2

wb.save(OUT)
print(f"Workbook saved: {OUT}")
